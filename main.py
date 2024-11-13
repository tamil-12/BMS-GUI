import tkinter as tk
from tkinter import Button, Entry, Frame, Label, StringVar, ttk, simpledialog, messagebox
import paho.mqtt.client as mqtt
import json
import threading
import pandas as pd
import os
import tkinter.font as font
from PIL import Image,ImageTk
from datetime import datetime
from openpyxl import Workbook,load_workbook

# Global variables
is_saving_to_excel = False
mqtt_data = []
data = {}
device_list = []
current_window = None
root = tk.Tk()
root.withdraw()  # Hide the root window initially
labels = {}
cell_labels = []
protection_labels = {}
# Define parameters at the global level to avoid scoping issues
parameters = [
    "Total Voltage", "Temperature Sensor 1", "Humidity", "Capacity Remaining",
    "Power", "Current", "Capacity Remaining (Ah)", "Nominal Capacity (Ah)", "MOSFET Charge"
]

protection_statuses = {
        "Single_Cell_Overvoltage_Count": None,
        "Single_Cell_Undervoltage_Count": None,
        "Whole_Pack_Overvoltage_Count": None,
        "Whole_Pack_Undervoltage_Count": None,
        "Charging_Over_Temperature_Count": None,
        "Charging_Low_Temperature_Count": None,
        "Discharge_Over_Temperature_Count": None,
        "Discharge_Low_Temperature_Count": None,
        "Charging_Overcurrent_Count": None,
        "Discharge_Overcurrent_Count": None,
        "Short_Circuit_Protection_Count": None,
        "Front_end_Detection_IC_Error_Count": None,
        "Software_Lock_MOS_Count": None
    }



device_file = "C:\\data_files\\devices.xlsx"
login_file = "C:\\data_files\\login_data.xlsx"
os.makedirs("C:\\data_files", exist_ok=True)

if os.path.exists(device_file):
    device_list = pd.read_excel(device_file)["Device_Name"].tolist()

if os.path.exists(login_file):
    login_data = pd.read_excel(login_file)
else:
    login_data = pd.DataFrame(columns=["Name", "Username", "Password"])





# MQTT callback
def on_message(client, userdata, message):
    global data
    global mqtt_data

    data = json.loads(message.payload.decode('utf-8'))
    mqtt_data.append(data)
   

    root.after(0, update_gui, data)
    if is_saving_to_excel:
        save_data_to_excel(data)  # Update the GUI with new data

def save_data_to_excel(data):
    # Define a fixed path for the Excel file in the Downloads folder
    downloads_folder = os.path.expanduser("~/Downloads")
    file_path = os.path.join(downloads_folder, "MQTT_Data.xlsx")  # Fixed file name

    # Check if the file exists
    if not os.path.exists(file_path):
        # Create a new workbook and add headers if the file doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "MQTT Data"  # Set a title for the sheet
        # Define column headers
        headers = ["Timestamp", "Total Voltage", "Temperature Sensor 1", "Humidity", "Capacity Remaining",
                   "Power", "Current", "Capacity Remaining (Ah)", "Nominal Capacity (Ah)", "MOSFET Charge"] \
                  + [f"Cell {i+1} Voltage" for i in range(13)] + list(protection_statuses.keys())
        sheet.append(headers)  # Append headers to the Excel sheet
    else:
        # Load the existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Select the active sheet

    # Prepare the data row with timestamp and the values from data dictionary
    row_data = [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]  # Timestamp
    # Append main data fields
    row_data += [
        data.get("Total_Voltage", "N/A"),
        data.get("Temperature_Sensor_1", "N/A"),
        data.get("Humidity", "N/A"),
        data.get("Capacity_Remaining_Percent", "N/A"),
        data.get("Watts", "N/A"),
        data.get("Amps", "N/A"),
        data.get("Capacity_Remaining_Ah", "N/A"),
        data.get("Nominal_Capacity_Ah", "N/A"),
        data.get("Mosfet_Charge", "N/A")
    ]
    # Append cell voltages
    cell_data = data.get("Cells", [])
    for i in range(13):  # Assuming 13 cells
        cell_voltage = cell_data[i].get("Voltage", "N/A") if i < len(cell_data) else "N/A"
        row_data.append(cell_voltage)
    # Append protection status values
    protection_data = data.get("Protection_Status", {})
    for key in protection_statuses.keys():
        row_data.append(protection_data.get(key, "N/A"))

    # Append the row data to the sheet
    sheet.append(row_data)

    # Save the workbook
    workbook.save(file_path)
def start_saving():
    global is_saving_to_excel
    is_saving_to_excel = True
    save_button.grid_remove()  # Hide the "Save to Excel" button
    stop_button.grid()         # Show the "Stop Saving" button

def stop_saving():
    global is_saving_to_excel
    is_saving_to_excel = False
    stop_button.grid_remove()  # Hide the "Stop Saving" button
    save_button.grid()  
def connect_mqtt(device_name=None):
    client = mqtt.Client()
    client.on_message = on_message
    client.connect("mqtt.eclipseprojects.io", 1883, 60)
    if device_name:
        payload = json.dumps({"device_name": device_name})
        client.publish("bms/connect", payload)

    client.subscribe("bms/data")
    client.loop_start()
def update_gui(data):
    # Update Basic Information labels
    labels["Total Voltage"].config(text=f"Total Voltage: {data.get('Total_Voltage', 'N/A')} V")
    labels["Temperature Sensor 1"].config(text=f"Temperature Sensor 1: {data.get('Temperature_Sensor_1', 'N/A')} °C")
    labels["Humidity"].config(text=f"Humidity: {data.get('Humidity', 'N/A')} %")
    labels["Capacity Remaining"].config(text=f"Capacity Remaining: {data.get('Capacity_Remaining_Percent', 'N/A')} %")
    labels["Power"].config(text=f"Power: {data.get('Watts', 'N/A')} W")
    labels["Current"].config(text=f"Current: {data.get('Amps', 'N/A')} A")
    labels["Capacity Remaining (Ah)"].config(text=f"Capacity Remaining (Ah): {data.get('Capacity_Remaining_Ah', 'N/A')} Ah")
    labels["Nominal Capacity (Ah)"].config(text=f"Nominal Capacity (Ah): {data.get('Nominal_Capacity_Ah', 'N/A')} Ah")
    labels["MOSFET Charge"].config(text=f"MOSFET Charge: {data.get('Mosfet_Charge', 'N/A')}")

    # Update Cell Voltages
    cell_data = data.get("Cells", [])  # Ensure "Cells" is a list of cell data
    for i, cell_label in enumerate(cell_labels):
        if i < len(cell_data):
            cell = cell_data[i]
            cell_label.config(text=f"Cell {i + 1} Voltage: {cell.get('Voltage', 'N/A')} V")  # Use 'Voltage' key
        else:
            cell_label.config(text=f"Cell {i + 1} Voltage: N/A")
    
    # Update Protection Status
   # Debug protection status updates
    protection_data = data.get("Protection_Status", {})
   
    for key, label in protection_labels.items():
       
        label.config(text=f"{key}: {protection_data.get(key, 'N/A')}")


# Admin Page
def open_admin_page():
    global current_window
    close_current_window()

    admin_page = tk.Toplevel(root)
    admin_page.title("Admin Page")
    admin_page.geometry("800x500")
    current_window = admin_page

    devices_frame = ttk.LabelFrame(admin_page, text="Devices", padding=(10, 5))
    devices_frame.pack(padx=10, pady=10, fill="x")

    def refresh_device_list():
        for widget in devices_frame.winfo_children():
            widget.destroy()

        for device in device_list:
            frame = ttk.Frame(devices_frame)
            frame.pack(anchor="w", fill="x", pady=2)
            device_label = ttk.Label(frame, text=device, padding=5)
            device_label.pack(side="left")
            remove_button = ttk.Button(frame, text="Remove", command=lambda d=device: remove_device(d))
            remove_button.pack(side="right")

    def add_device():
        device_name = simpledialog.askstring("Add Device", "Enter device name:")
        if device_name and device_name not in device_list:
            device_list.append(device_name)
            pd.DataFrame(device_list, columns=["Device_Name"]).to_excel(device_file, index=False)
            messagebox.showinfo("Success", f"Device '{device_name}' added successfully!")
            refresh_device_list()

    def remove_device(device_name):
        if device_name in device_list:
            device_list.remove(device_name)
            pd.DataFrame(device_list, columns=["Device_Name"]).to_excel(device_file, index=False)
            messagebox.showinfo("Success", f"Device '{device_name}' removed successfully!")
            refresh_device_list()

    refresh_device_list()
    ttk.Button(admin_page, text="Add Device", command=add_device).pack(pady=5)
    ttk.Button(admin_page, text="Back", command=open_login_page).pack(pady=10)

# Device Selection Page
def open_device_selection_page():
    global current_window
    close_current_window()

    device_selection_page = tk.Toplevel(root)
    device_selection_page.title("Select Device")
    device_selection_page.geometry("800x500")
    current_window = device_selection_page

    for device in device_list:
        frame = ttk.Frame(device_selection_page)
        frame.pack(padx=10, pady=5, fill="x")
        device_label = ttk.Label(frame, text=device)
        device_label.pack(side="left")
        connect_button = ttk.Button(frame, text="Connect", command=lambda d=device: connect_and_open_main_page(d))
        connect_button.pack(side="right")

    ttk.Button(device_selection_page, text="Back", command=open_login_page).pack(pady=10)

# Connect and Open Main Page
def connect_and_open_main_page(device_name):
    connect_mqtt(device_name)
    open_main_app()
# Login Function

def login(username_entry, password_entry):
    username = username_entry.get()
    password = password_entry.get()

    if username == "BMS1" and password == "12345":
        open_admin_page()
    else:
        # Ensure username and password are treated as strings
        login_data['Username'] = login_data['Username'].astype(str)
        login_data['Password'] = login_data['Password'].astype(str)

        # Validate user credentials from Excel
        user = login_data[(login_data['Username'] == str(username)) & (login_data['Password'] == str(password))]
        
        if not user.empty:
            open_device_selection_page()
        else:
            error_label.config(text="Invalid username or password", foreground="red")


# Create User Function
def create_user():
    def save_user():
        name1 = name.get()
        username = user_name.get()
        password = pas.get()

        if name and username and password:
            new_user = pd.DataFrame([[name1, username, password]], columns=["Name", "Username", "Password"])
            global login_data
            login_data = pd.concat([login_data, new_user], ignore_index=True)
            login_data.to_excel(login_file, index=False)
            messagebox.showinfo("Success", "User created successfully!")
            create_user_window.destroy()
        else:
            messagebox.showwarning("Incomplete Data", "Please fill in all fields.")

    # Create User Window
    create_user_window = tk.Toplevel(root)
    create_user_window.geometry("800x500")
    create_user_window.title("Create New User")
    create_user_window.config(bg="white")
    file_path = os.path.join(os.path.dirname(__file__), "login_image.jpg")
    img_path = Image.open(file_path)
    image_resize = img_path.resize((350, 400))
    img = ImageTk.PhotoImage(image_resize)
    Label(create_user_window, image=img).place(x=30, y=40)
    create_user_window.image = img  # Keep a reference to avoid garbage collection

    # Right-side frame for the login fields
    frame1 = Frame(create_user_window, width=350, height=400, bg="white")
    frame1.place(x=415, y=40)

 
   
    name = StringVar()
    name = Entry(create_user_window, width=30, font=("Microsoft JhengHei", 10), border=0, fg="#41484a", textvariable=name)
    name.insert(0, "Name")  # Placeholder
    name.place(x=470, y=130)
    name.bind('<FocusIn>', lambda e: on_enter(name, "Name"))
    name.bind('<FocusOut>', lambda e: on_leave(name, "Name"))
    Frame(create_user_window, height=2, width=290, bg="black").place(x=470, y=150)

    user_name = StringVar()
    user_name = Entry(create_user_window, width=30, font=("Microsoft JhengHei", 10), border=0, fg="#41484a", textvariable=user_name)
    user_name.insert(0, "Username")  # Placeholder
    user_name.place(x=470, y=180)
    user_name.bind('<FocusIn>', lambda e: on_enter(user_name, "Username"))
    user_name.bind('<FocusOut>', lambda e: on_leave(user_name, "Username"))
    Frame(create_user_window, height=2, width=290, bg="black").place(x=470, y=200)

    # Password Entry
    pas = StringVar()
    password = Entry(create_user_window, width=30, font=("Microsoft JhengHei", 10), border=0, fg="#41484a", textvariable=pas, show="*")
    password.insert(0, "Password")  # Placeholder
    password.place(x=470, y=230)
    password.bind('<FocusIn>', lambda e: on_enter(password, "Password"))
    password.bind('<FocusOut>', lambda e: on_leave(password, "Password"))
    Frame(create_user_window, height=2, width=290, bg="black").place(x=470, y=250)
    signin_button = Button(create_user_window, text='Create', fg="white", bg="#0eb1ed", width=35, border=0, height=2, font=("Microsoft JhengHei", 8, "bold"), command=save_user)
    signin_button.place(x=485, y=290)

class ScrollableFrame(ttk.Frame):
    """ A frame that can be scrolled. """
    def __init__(self, parent):
        super().__init__(parent)
        self.canvas = tk.Canvas(self)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        # Configure the scrollable frame
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Place the canvas and scrollbar in the grid
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.scrollbar.grid(row=0, column=1, sticky="ns")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Configure the row and column of the scrollable frame to expand
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

def open_main_app():
    global root, current_window, protection_labels,save_button, stop_button
    close_current_window()
    root = tk.Toplevel()
    root.title("Battery Management System")
    root.geometry("900x600")  # Set a fixed size for the initial window
    current_window = root

    # Make the root window resizable
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Create scrollable frame
    scrollable_frame = ScrollableFrame(root)
    scrollable_frame.grid(row=0, column=0, sticky="nsew")

    # Define a larger font for readability
    large_font = font.Font(size=10)

    # Main frame to contain everything with padding
    main_frame = ttk.Frame(scrollable_frame.scrollable_frame, padding=(10, 10))
    main_frame.grid(row=0, column=0, sticky="nsew")

    # Configure the main frame to expand
    main_frame.grid_columnconfigure(0, weight=1)
    main_frame.grid_columnconfigure(1, weight=1)
    main_frame.grid_columnconfigure(2, weight=1)

    # Function to create a labeled frame for each parameter with default "N/A" and black border
    def create_param_box(parent, label_text):
        box_frame = ttk.LabelFrame(parent, padding=(5, 2), style="BlackBorder.TLabelframe")
        label = ttk.Label(box_frame, text="N/A", font=large_font)
        label.pack(padx=5, pady=5, expand=True)
        return box_frame, label

    # Create and configure Basic Information frame
    parameters_frame = ttk.LabelFrame(main_frame, text="Basic Information", padding=(10, 5), style="BlackBorder.TLabelframe")
    parameters_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    parameters_frame.grid_columnconfigure(0, weight=1)

    # Add parameters in individual boxes with "N/A" as default
    for i, param in enumerate(parameters):
        box, label = create_param_box(parameters_frame, param)
        box.grid(row=i, column=0, padx=5, pady=5, sticky="ew")
        labels[param] = label

    # Create and configure Cell Voltages frame
    cells_frame = ttk.LabelFrame(main_frame, text="Cell Voltages", padding=(10, 5), style="BlackBorder.TLabelframe")
    cells_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
    cells_frame.grid_columnconfigure(0, weight=1)

    # Create a box for each cell voltage with "N/A" as default (assuming 13 cells)
    for i in range(13):
        box, label = create_param_box(cells_frame, f"Cell {i+1} Voltage")
        box.grid(row=i, column=0, padx=5, pady=5, sticky="ew")
        cell_labels.append(label)

    # Create and configure Protection Status frame
    protection_frame = ttk.LabelFrame(main_frame, text="Protection Status", padding=(10, 5), style="BlackBorder.TLabelframe")
    protection_frame.grid(row=0, column=2, padx=10, pady=10, sticky="nsew")
    protection_frame.grid_columnconfigure(0, weight=1)

    # Populate the protection labels dictionary
    for i, protection in enumerate(protection_statuses):
        box, label = create_param_box(protection_frame, protection)
        box.grid(row=i, column=0, padx=5, pady=5, sticky="ew")
        protection_labels[protection] = label

   
    # Initialize Save and Stop Saving buttons but only display the "Save to Excel" button initially
    save_button = ttk.Button(main_frame, text="Save to Excel", command=start_saving)
    stop_button = ttk.Button(main_frame, text="Stop Saving", command=stop_saving)
    save_button.grid(row=1, column=1, pady=10, sticky="ew")  # Display "Save to Excel" button initially
    stop_button.grid(row=1, column=1, pady=10, sticky="ew")
    stop_button.grid_remove()  # Hide "Stop Saving" initially


    # Add "Back" button at the bottom of main_frame, spanning across all columns
    back_button = ttk.Button(main_frame, text="Back", command=open_device_selection_page)
    back_button.grid(row=3, column=0, columnspan=3, pady=10)

    # Configure style for padding, font size, and black border

def on_enter(entry, placeholder):
    """Clears the placeholder text when the entry is focused."""
    if entry.get() == placeholder:
        entry.delete(0, 'end')  # Clear the entry

def on_leave(entry, placeholder):
    """Restores the placeholder text if the entry is empty."""
    if entry.get() == "":
        entry.insert(0, placeholder)  # Restore placeholder
# Login Page
def open_login_page():
    global current_window,error_label
    close_current_window()

    # Main login window
    main_frame_window = tk.Toplevel(root)
    main_frame_window.title("BMS")
    main_frame_window.geometry("800x500")
    main_frame_window.config(bg="white")
    main_frame_window.resizable(False, False)
    current_window = main_frame_window

    # Load and place the login image
    file_path = os.path.join(os.path.dirname(__file__), "login_image.jpg")
    img_path = Image.open(file_path)
    image_resize = img_path.resize((350, 400))
    img = ImageTk.PhotoImage(image_resize)
    Label(main_frame_window, image=img).place(x=30, y=40)
    main_frame_window.image = img  # Keep a reference to avoid garbage collection

    # Right-side frame for the login fields
    frame1 = Frame(main_frame_window, width=350, height=400, bg="white")
    frame1.place(x=415, y=40)

    signin_label = Label(main_frame_window, text="Sign-In", bg="white", fg="#05739c", font=("Microsoft JhengHei", 23, "bold"))
    signin_label.place(x=525, y=70)

    signin_name = StringVar()
    sign_user_name = Entry(main_frame_window, width=30, font=("Microsoft JhengHei", 10), border=0, fg="#41484a", textvariable=signin_name)
    sign_user_name.insert(0, "Username")  # Placeholder
    sign_user_name.place(x=470, y=180)
    sign_user_name.bind('<FocusIn>', lambda e: on_enter(sign_user_name, "Username"))
    sign_user_name.bind('<FocusOut>', lambda e: on_leave(sign_user_name, "Username"))
    Frame(main_frame_window, height=2, width=290, bg="black").place(x=470, y=200)

    # Password Entry
    sign_pas = StringVar()
    sign_password = Entry(main_frame_window, width=30, font=("Microsoft JhengHei", 10), border=0, fg="#41484a", textvariable=sign_pas, show="*")
    sign_password.insert(0, "Password")  # Placeholder
    sign_password.place(x=470, y=230)
    sign_password.bind('<FocusIn>', lambda e: on_enter(sign_password, "Password"))
    sign_password.bind('<FocusOut>', lambda e: on_leave(sign_password, "Password"))
    Frame(main_frame_window, height=2, width=290, bg="black").place(x=470, y=250)
    # Sign-in button
    signin_button = Button(main_frame_window, text='Sign-in', fg="white", bg="#0eb1ed", width=35, border=0, height=2, font=("Microsoft JhengHei", 8, "bold"), command=lambda: login(sign_user_name, sign_password))
    signin_button.place(x=485, y=290)
    error_label = ttk.Label(main_frame_window, text="", font=("Arial", 10), foreground="red")
    error_label.pack()

    # Create User button
    create_user_button = Button(main_frame_window, text="Create User", fg="white", bg="#05739c", width=35, border=0, height=2, font=("Microsoft JhengHei", 8, "bold"), command=create_user)
    create_user_button.place(x=485, y=340)

# Close Current Window
def close_current_window():
    global current_window
    if current_window:
        current_window.withdraw()
        current_window = None

# Main Program Execution
open_login_page()
root.mainloop()
